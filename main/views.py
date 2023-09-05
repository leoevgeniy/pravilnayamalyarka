import openpyxl
import xlrd
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
import json
from datetime import date
from math import ceil

from openpyxl.styles import Border, Side, Font, Alignment

from cms.forms import UploadFileForm, SearchForm
from cms.models import Product, Service, PromoSlider, WorkPhoto, Logo, Introduction, Socials, Contacts, Packprice, \
    Services_files, OurPhoto, Services_calculation_cost
from crm.models import Order, StatusCrm, OrderItems
from .models import Category, SubCategory, ServiceCategory
from crm.forms import OrderForm
from telebot.sendmessage import send_telegram
from PIL import Image
from pravilnayamalyarka.settings import BASE_DIR
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from meta.views import Meta
from urllib.parse import unquote
import xlwt


# Create your views here.

def index(request):
    today = date.today()
    promos = PromoSlider.objects.filter(start_date__lte=today).filter(expiration_date__gte=today)
    work = WorkPhoto.objects.all()

    services_category = ServiceCategory.objects.all()
    work = WorkPhoto.objects.all()
    work_landscape = []
    work_portrate = []
    for w in work:
        with Image.open(BASE_DIR + w.photo_url) as img:
            width, height = img.size
            if width / height >= 1.77 and len(work_landscape) < 11:
                work_landscape.append(w)
            elif width / height < 1 and len(work_portrate) < 11:
                work_portrate.append(w)
    categories = Category.objects.all()
    subcategories = SubCategory.objects.all()
    form = OrderForm
    searchform = SearchForm
    socials = Socials.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        intro = Introduction.objects.get(inuse=True)
    except:
        intro = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''
    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()
    meta = Meta(
        title="Правильная малярка",
        description='Только проверенные поставщики, занимающие лидирующие позиции на мировом рынке. Качество подтверждено мировыми гигантами.',
        keywords=['pony', 'ponies', 'awesome'],
        extra_props={
            'viewport': 'width=device-width, initial-scale=1.0, minimum-scale=1.0'
        },
        extra_custom_props=[
            ('http-equiv', 'Content-Type', 'text/html; charset=UTF-8'),
        ]
    )
    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))

    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'cart_products': ids,
        'pagename': 'index',
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'work': work,
        'contacts': contacts[0],
        'socials': socials,
        'logo': logo,
        'intro': intro,
        'work_landscape': work_landscape,
        'work_portrate': work_portrate,
        'categories': categories,
        'subcategories': subcategories,
        'promoslider': promos,
        'services_category': services_category,
        'form': form,
        'searchform': searchform,
    }
    return render(request, 'main/index.html', disc)


def about(request):
    pricelist = Services_files.objects.all()[0]
    work = OurPhoto.objects.all()
    work_landscape = []
    work_portrate = []
    for w in work:
        with Image.open(BASE_DIR + w.photo_url) as img:
            width, height = img.size
            if width / height >= 1.77:
                work_landscape.append(w)
            elif width / height < 1:
                work_portrate.append(w)
    form = OrderForm
    searchform = SearchForm
    socials = Socials.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        intro = Introduction.objects.get(inuse=True)
    except:
        intro = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'cart_products': ids,
        'pricelist': pricelist,
        'pagename': 'about',
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'contacts': contacts[0],
        'socials': socials,
        'logo': logo,
        'intro': intro,
        'work': work,
        'work_landscape': work_landscape,
        'work_portrate': work_portrate,
        'form': form,
        'searchform': searchform,
    }
    return render(request, 'main/about.html', disc)


def contacts(request):
    work = WorkPhoto.objects.all()
    work_landscape = []
    work_portrate = []
    for w in work:
        with Image.open(BASE_DIR + w.photo_url) as img:
            width, height = img.size
            if width / height >= 1.77:
                work_landscape.append(w)
            elif width / height < 1:
                work_portrate.append(w)
    form = OrderForm
    searchform = SearchForm
    socials = Socials.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        intro = Introduction.objects.get(inuse=True)
    except:
        intro = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'cart_products': ids,
        'pagename': 'contacts',
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'contacts': contacts[0],
        'socials': socials,
        'logo': logo,
        'intro': intro,
        'work_landscape': work_landscape,
        'work_portrate': work_portrate,
        'form': form,
        'searchform': searchform,
    }
    return render(request, 'main/contacts.html', disc)


def delivery(request):
    work = WorkPhoto.objects.all()
    work_landscape = []
    work_portrate = []
    for w in work:
        with Image.open(BASE_DIR + w.photo_url) as img:
            width, height = img.size
            if width / height >= 1.77:
                work_landscape.append(w)
            elif width / height < 1:
                work_portrate.append(w)
    form = OrderForm
    searchform = SearchForm
    socials = Socials.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        intro = Introduction.objects.get(inuse=True)
    except:
        intro = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'cart_products': ids,
        'pagename': 'delivery',
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'contacts': contacts[0],
        'socials': socials,
        'logo': logo,
        'intro': intro,
        'work_landscape': work_landscape,
        'work_portrate': work_portrate,
        'form': form,
        'searchform': searchform,
    }
    return render(request, 'main/delivery.html', disc)


def goods(request):
    catalog = {}
    categories = Category.objects.all()
    form = OrderForm
    up = UploadFileForm
    searchform = SearchForm
    for category in categories:
        subCategory = SubCategory.objects.filter(category=category)
        subcat = []
        for sub in subCategory:
            subcat.append(sub.name)
        catalog.update({category.name: subcat})
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    disc = {
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'contacts': contacts[0],
        'logo': logo,
        'searchform': searchform,
        'catalog': catalog,
        'form': form,
        'up': up,
    }

    return render(request, 'main/goods.html', disc)


def services(request):
    work = WorkPhoto.objects.all()

    catalog = {}
    categories = ServiceCategory.objects.all()
    searchform = SearchForm
    # form = OrderForm
    # up = UploadFileForm
    for category in categories:
        services = Service.objects.filter(service_category=category)
        currentservices = []
        for ser in services:
            currentservices.append(ser)
        catalog.update({category.name: currentservices})
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        download = Services_files.objects.get(inuse=True)
    except:
        download = ''
    try:
        calc = Services_calculation_cost.objects.get(inuse=True)
    except:
        calc = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''
    socials = Socials.objects.all()
    form = OrderForm

    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()
    # pricelist = Services_files.objects.all()[0]
    disc = {
        'calc': calc,
        'download': download,
        # 'pricelist': pricelist,
        'categories': categories,
        'work': work,
        'pagename': 'services',
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'form': form,
        'socials': socials,
        'contacts': contacts[0],
        'logo': logo,
        'searchform': searchform,
        'catalog': catalog,
    }

    return render(request, 'main/services.html', disc)


def category(request, category):
    query = ''

    if request.GET.get('text') is not None:
        print(request.GET.get('text'))
        query = request.GET.get('text')

    try:
        after_buy = request.GET.get('cart')
    except:
        after_buy = 0
    category = Category.objects.get(name=unquote(category))
    subcat_name = ''
    if request.GET.get('subcategory') is not None:
        subcat_name = (request.GET.get('subcategory'))
        subcat = SubCategory.objects.get(category=category, name__exact=subcat_name)
    else:
        subcat = SubCategory.objects.filter(category=category)

    today = date.today()
    promos = PromoSlider.objects.filter(start_date__lte=today).filter(expiration_date__gte=today)

    subcategory = SubCategory.objects.filter(category=category)
    prods2 = Product.objects.filter(category=category, subcategory__exact=None, name__iregex=query.lower())
    products = []
    for pr in prods2:
        products.append(pr)
    for i in range(0, len(products)):
        for j in range(i+1, len(products)):
            if Packprice.objects.filter(product=products[j])[0].price > Packprice.objects.filter(product=products[i])[0].price:
                products[j], products[i] = products[i], products[j]

    productswithoutquery = Product.objects.filter(category=category, subcategory__exact=None)

    try:
        products_firstsubcat = Product.objects.filter(category=category, subcategory=subcat[0])
    except:
        products_firstsubcat = []
    searchform = SearchForm
    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''

    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    socials = Socials.objects.all()

    pageinput = request.GET.get('page')
    page = request.GET.get('page')

    paginator = Paginator(products, 10000)
    try:
        products = paginator.page(page)
        products_firstsubcat = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
        products_firstsubcat = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)
        products_firstsubcat = paginator.page(paginator.num_pages)

    if page == None or page == 'undefined':
        page = 1

    page = int(page)
    pages = []
    for p in range(paginator.num_pages):
        pages.append(p + 1)
    form = OrderForm
    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'productswithoutquery':productswithoutquery,
        'query': query,
        'cart_products': ids,
        "products_first": products_firstsubcat,
        'after_buy': after_buy,
        'promoslider': promos,
        'pagename': 'category',
        'form': form,
        'pageinput': pageinput,
        'page': page,
        'pages': pages,
        "products": products,

        'socials': socials,
        'contacts': contacts[0],

        'logo': logo,
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'searchform': searchform,
        'category': category,
        'subcategory': subcategory
    }
    return render(request, 'main/category.html', disc)


def subcategory(request, category, subcategory, *args):

    vendor = ''
    sortup = ''
    sortdown = ''
    query = ''

    if request.GET.get('text') is not None:
        query = request.GET.get('text')
    if request.GET.get('vendor') is not None:
        vendor = (request.GET.get('vendor'))
    if request.GET.get('sortup') is not None:
        sortup = (request.GET.get('sortup'))
    if request.GET.get('sortdown') is not None:
        sortdown = (request.GET.get('sortdown'))

    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    try:
        category = Category.objects.get(name=unquote(category))
    except:
        category = ''
    try:
        subCategory = SubCategory.objects.get(category=category, name=unquote(subcategory))
    except:
        subCategory = ''
    try:
        this_subcategories = SubCategory.objects.filter(category=category)
    except:
        this_subcategories = ''
    searchform = SearchForm
    if sortup:
        prods2 = Product.objects.filter(category=category, subcategory=subCategory, name__iregex=query.lower())
        products = []
        vc = []
        prods1 = []
        for pr in prods2:
            prods1.append(pr)
        for i in range(0, len(prods1)):
            for j in range(i+1, len(prods1)):
                if Packprice.objects.filter(product=prods1[j])[0].price < Packprice.objects.filter(product=prods1[i])[0].price:
                    prods1[j], prods1[i] = prods1[i], prods1[j]

        for pr in prods1:
                # vc = pr.vendor_code
                if pr.vendor_code not in vc:
                    products.append(pr)
                    # print(vc)
                    vc.append(pr.vendor_code)
    else:
        prods2 = Product.objects.filter(category=category, subcategory=subCategory, name__iregex=query.lower())
        prods1 = []
        for pr in prods2:
            prods1.append(pr)
        for i in range(0, len(prods1)):
            for j in range(i+1, len(prods1)):
                try:
                    if Packprice.objects.filter(product=prods1[j])[0].price > Packprice.objects.filter(product=prods1[i])[0].price:
                        prods1[j], prods1[i] = prods1[i], prods1[j]
                except:
                    pass


        products = []
        vc = []
        for pr in prods1:
            # vc = pr.vendor_code
            if pr.vendor_code not in vc:
                products.append(pr)
                # print(vc)
                vc.append(pr.vendor_code)


    allbrend = []
    for product in Product.objects.filter(category=category, subcategory=subCategory):
        if product.vendor not in allbrend:
            allbrend.append(product.vendor)
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''

    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    socials = Socials.objects.all()
    pageinput = request.GET.get('page')
    page = request.GET.get('page')

    paginator = Paginator(products, 10000)
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    if page == None or page == 'undefined':
        page = 1

    page = int(page)
    pages = []
    for p in range(paginator.num_pages):
        pages.append(p + 1)
    form = OrderForm
    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'query': query,
        'cart_products': ids,
        'form': form,
        'pageinput': pageinput,
        'page': page,
        'pages': pages,
        'socials': socials,
        'contacts': contacts[0],

        'logo': logo,
        'searchform': searchform,
        'sortup': sortup,
        'sortdown': sortdown,
        'vendor': vendor,
        'allbrend': allbrend,
        'category': category,
        'subcategory': subCategory,
        'this_subcategory': this_subcategories,
        'products': products,
        # 'productsJS': productsJSData,
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
    }
    return render(request, 'main/subcategory.html', disc)


def thanks_page(request):
    searchform = SearchForm
    today = date.today()

    promos = PromoSlider.objects.filter(start_date__lte=today).filter(expiration_date__gte=today)

    name = request.POST['name']
    phone = request.POST['phone']
    status = StatusCrm.objects.get(status_name__exact='Новая')
    element = Order(order_name=name, order_phone=phone, order_type='Заказ обратного звонка', order_status=status)
    element.save()
    send_telegram(name, phone)
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    socials = Socials.objects.all()
    categories = Category.objects.all()

    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    form = OrderForm
    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'promoslider': promos,

        'categories': categories,
        'cart_products': ids,
        'form': form,
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        'socials': socials,
        'contacts': contacts[0],

        'name': name,
        'searchform': searchform,
        'logo': logo,
    }

    return render(request, 'main/index.html', disc)


def createbill(request):
    orderitems = OrderItems.objects.filter(order=int(request.GET.get('order_id')))
    # workbook = xlrd.open_workbook(BASE_DIR + '/static/billtamplate.xlsx', on_demand=True)
    pxl_doc = openpyxl.load_workbook(BASE_DIR + '/static/billtamplate.xlsx', data_only=True)
    sheet = pxl_doc.active
    thins = Side(border_style="thin", color="000000")
    medium = Side(border_style="thick", color="000000")
    double = Side(border_style="dashDot", color="ff0000")
    border = Border(
        left=Side(border_style=None, color='FF000000'),
        right=Side(border_style=None, color='FF000000'),
        top=Side(border_style=None, color='FF000000'),
        bottom=Side(border_style=None, color='FF000000'),
        diagonal=Side(border_style=None, color='FF000000'),
        diagonal_direction=0,
        outline=Side(border_style=None, color='FF000000'),
        vertical=Side(border_style=None, color='FF000000'),
        horizontal=Side(border_style=None, color='FF000000')
    )
    alignment = Alignment(
        horizontal='left',
        wrap_text=True,
        shrink_to_fit=True
    )
    alignmentNDS = Alignment(
        horizontal='right',
        wrap_text=True,
        shrink_to_fit=True
    )

    font = Font(name='Calibri',
        size = 11,
        bold = False,
        italic = False,
        vertAlign = None,
        underline = 'none',
        strike = False,
        color = 'FF000000')
    row = 18
    sheet.insert_rows(row, len(orderitems))
    count_row = 0
    total_cost = 0
    for item in orderitems:
        total_cost += item.price * item.qty
        try:
            sheet.unmerge_cells('A' + str(row + count_row) + ':' + 'F' + str(row + count_row))
        except:
            pass
        for cell in ['A', 'B', 'C', 'D', 'E', 'F']:
            if cell == 'A':
                if count_row + 1 == len(orderitems):
                    sheet[cell + str(row + count_row)].border = Border(left=medium, bottom=medium, right=thins)
                else:
                    sheet[cell + str(row + count_row)].border = Border(left=medium, right=thins)
                sheet[cell + str(row + count_row)] = count_row + 1
                sheet[cell + str(row + count_row)].font = font

            elif cell == 'B':
                if count_row + 1 == len(orderitems):
                    sheet[cell + str(row + count_row)].border = Border(bottom=medium, right=thins)
                else:
                    sheet[cell + str(row + count_row)].border = Border(right=thins)

                sheet[cell + str(row + count_row)] = item.product.name
                sheet[cell + str(row + count_row)].font = font
                sheet[cell + str(row + count_row)].alignment = alignment
                words_count_at_one_row = sheet.column_dimensions[cell].width / 1
                lines = ceil(len(str(sheet[cell + str(row + count_row)].value)) / words_count_at_one_row)
                sheet.row_dimensions[row + count_row].height = max(16, lines * 16)
            elif cell == 'C':
                if count_row + 1 == len(orderitems):
                    sheet[cell + str(row + count_row)].border = Border(bottom=medium, right=thins)
                else:
                    sheet[cell + str(row + count_row)].border = Border(right=thins)

                sheet[cell + str(row + count_row)] = item.qty
                sheet[cell + str(row + count_row)].font = font
            elif cell == 'D':
                if count_row + 1 == len(orderitems):
                    sheet[cell + str(row + count_row)].border = Border(bottom=medium, right=thins)
                else:
                    sheet[cell + str(row + count_row)].border = Border(right=thins)

                sheet[cell + str(row + count_row)] = item.product.unitofmeasure
                sheet[cell + str(row + count_row)].font = font
            elif cell == 'E':
                if count_row + 1 == len(orderitems):
                    sheet[cell + str(row + count_row)].border = Border(bottom=medium, right=thins)
                else:
                    sheet[cell + str(row + count_row)].border = Border(right=thins)

                sheet[cell + str(row + count_row)] = item.price
                sheet[cell + str(row + count_row)].font = font
            elif cell == "F":
                if count_row + 1 == len(orderitems):
                    sheet[cell + str(row + count_row)].border = Border(bottom=medium, right=medium)
                else:
                    sheet[cell + str(row + count_row)].border = Border(right=medium)

                sheet[cell + str(row + count_row)] = item.price * item.qty
                sheet[cell + str(row + count_row)].font = font
        # sheet.row_dimensions[row+count_row]
        count_row += 1

        try:
            sheet.unmerge_cells('A' + str(row + count_row) + ':' + 'F' + str(row + count_row))
        except:
            pass
        try:
            sheet.unmerge_cells('A' + str(row + count_row + 1) + ':' + 'F' + str(row + count_row + 1))
        except:
            pass

        bold = Font(bold=True)
        sheet['E' + str(row + count_row)] = 'Итого:'
        # sheet['E' + str(row + count_row)].font.bold = True
        sheet['F' + str(row + count_row)] = total_cost
        sheet['F' + str(row + count_row)].font = bold
        sheet['F' + str(row + count_row+1)] = 'Без налога(НДС)'
        sheet['F' + str(row + count_row+1)].font = bold
        sheet['F' + str(row + count_row+1)].alignment = alignmentNDS
    # for i in range(len(list(sheet.rows))):
    #     if sheet['E'+str(i+1)].value == 'Итого:':
    #         sheet['F'+str(i+1)] = total_cost
    #         break

    pxl_doc.save(BASE_DIR + '/media/order' + request.GET.get('order_id') + '.xlsx')
    data = open(BASE_DIR + '/media/order' + request.GET.get('order_id') + '.xlsx', "br").read()

    response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=bill' + str(request.GET.get('order_id')) + '.xlsx'
    return response
    # return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
