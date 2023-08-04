import json
import os
from datetime import date
from PIL import Image
import xlrd
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponseRedirect
from django.shortcuts import render
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from urllib.parse import unquote
# Create your views here.
from cms.forms import UploadFileForm, SearchForm
from cms.models import WorkPhoto, Product, Service, Vendor, Packprice, Logo, Contacts, Socials, Introduction, \
    PromoSlider
from crm.forms import OrderForm
from main.models import SubCategory, Category, ServiceCategory
from pravilnayamalyarka.settings import MEDIAS_DIR, MEDIA_ROOT


def legal(request):
    today = date.today()
    promos = PromoSlider.objects.filter(start_date__lte=today).filter(expiration_date__gte=today)
    services_category = ServiceCategory.objects.all()
    work = WorkPhoto.objects.all()
    work_landscape = []
    work_portrate = []
    for w in work:
        with Image.open(MEDIAS_DIR + w.photo_url) as img:
            width, height = img.size
            if width / height >= 1.77 and len(work_landscape) < 11:
                work_landscape.append(w)
            elif width / height < 1 and len(work_portrate) < 11:
                work_portrate.append(w)
    categories = Category.objects.all()
    subcategories = SubCategory.objects.all()
    form = OrderForm
    searchform = SearchForm
    socials = Socials.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        intro = Introduction.objects.get(inuse=True)
    except:
        intro = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''
    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'cart_products': ids,
        'pagename': 'index',
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,

        'contacts': contacts[0],
        'socials': socials,
        'logo': logo,
        'intro': intro,
        'work_landscape': work_landscape,
        'work_portrate': work_portrate,
        'categories': categories,
        'subcategories': subcategories,
        'promoslider': promos,
        'services_category': services_category,
        'form': form,
        'searchform': searchform,
    }
    return render(request, 'main/policy.html', disc)


def product_view(request, pk):
    try:
        defaultweight = int(request.GET.get('weight'))
    except:
        defaultweight = ''
    searchform = SearchForm
    product = Product.objects.get(vendor_code=pk)
    # session_key = request.session.session_key
    # if not session_key:
    #     request.session["session_key"] = 123
    #     request.session.cycle_key()
    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''

    socials = Socials.objects.all()
    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))
    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass

    disc = {
        'defaultweight': defaultweight,
        'cart_products': ids,
        'pagename': 'product_details',
        'socials': socials,
        'contacts': contacts[0],
        'logo': logo,
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        "product": product,
        'searchform': searchform,
    }
    return render(request, 'main/product_details.html', disc)


def photomigrations(request):
    col_dir = BASE_DIR + '/materials/photo/'
    col = os.listdir(col_dir)
    for img in col:
        work = WorkPhoto.objects.create(
            name='',
            photo='images/donework/' + img,
        )

    return render(request, 'main/thanks_page.html')


def productsimport(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            filehandle = request.FILES.get('file', False)

            pxl_doc = openpyxl.load_workbook(filehandle.temporary_file_path())
            sheet = pxl_doc['Прайс']
            workbook = xlrd.open_workbook(filehandle.temporary_file_path())
            worksheet = workbook.sheet_by_index(0)
            rows = []
            image_loader = SheetImageLoader(sheet)
            cat = ''
            image = ''
            for row in range(1, worksheet.nrows):
                line = {}
                if type(worksheet.cell(row, 0).value) == str and not worksheet.cell(row, 1).value:
                    cat = worksheet.cell(row, 0).value
                if type(worksheet.cell(row, 0).value) == float:
                    params = []
                    for column in range(worksheet.ncols):
                        params.append(worksheet.cell(row, column).value)
                    params.append(cat)
                    try:

                        image = image_loader.get('C' + str(row + 1))
                    except:
                        pass
                    line[worksheet.cell(row, 0).value] = params
                    rows.append(line)
                    # print(worksheet.cell(row, 0).value, line.keys())
                    if Product.objects.filter(vendor_code=params[0]).exists():
                        product = Product.objects.get(vendor_code=int(params[0]))
                        image.save(MEDIA_ROOT + '/images/' + str(params[0]) + '.png', format='png')
                    else:
                        image.save(MEDIA_ROOT + '/images/' + str(params[0]) + '.png', format='png')
                        category = Category.objects.get(name='Ручной инструмент')
                        try:
                            subcategory = SubCategory.objects.get(name=cat)
                        except:
                            subcategory = SubCategory.objects.create(
                                category=category,
                                name=cat,
                                description='',

                            )

                        try:
                            vendor = Vendor.objects.get(name='ROLLINGDOG')
                        except:
                            vendor = Vendor.objects.create(name='ROLLINGDOG')
                        product = Product.objects.create(
                            category=category,
                            subcategory=subcategory,
                            photo='images/' + str(params[0]) + '.png',
                            name=str(params[1]),
                            description=str(params[3]),
                            vendor=vendor,
                            vendor_code=int(params[0]),
                            rrc=int(params[5]),
                            availability=str(params[6]),
                            pack=str(params[7])
                        )
                        pack = Packprice.objects.create(
                            product=product,
                            weight=1,
                            price=int(params[5]),
                            oldprice=int(params[5]) * 1.2
                        )

            # if Product.objects.get(vendor_code=)

        return HttpResponseRedirect('/admin/')


def services_import(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            filehandle = request.FILES.get('file', False)

            pxl_doc = openpyxl.load_workbook(filehandle.temporary_file_path())
            sheet = pxl_doc['Лист1']
            workbook = xlrd.open_workbook(filehandle.temporary_file_path())
            worksheet = workbook.sheet_by_index(0)
            rows = []
            cat = ''
            for row in range(1, worksheet.nrows):
                line = {}
                if type(worksheet.cell(row, 2).value) == str and not worksheet.cell(row, 1).value:
                    cat = worksheet.cell(row, 2).value

                if type(worksheet.cell(row, 1).value) == float:
                    params = []
                    for column in range(worksheet.ncols):
                        if column > 0:
                            params.append(worksheet.cell(row, column).value)
                    params.append(cat)
                    line[worksheet.cell(row, 1).value] = params
                    rows.append(line)
                    # print(worksheet.cell(row, 0).value, line.keys())
                    if Service.objects.filter(service_code=str(params[0])).exists():
                        service = Service.objects.get(service_code=str(params[0]))
                    else:
                        if ServiceCategory.objects.filter(name=cat).exists:
                            category = ServiceCategory.objects.get(name=cat)
                            service = Service.objects.create(
                                name=params[1],
                                service_pc=params[2],
                                service_price=params[5],
                                service_code=params[0],
                                service_category=category

                            )
                # if Product.objects.get(vendor_code=)

        return HttpResponseRedirect('/admin/')


def search(request):
    # today = date.today()
    # promos = PromoSlider.objects.filter(start_date__lte=today).filter(expiration_date__gte=today)
    # work = WorkPhoto.objects.all()

    # services_category = ServiceCategory.objects.all()
    # work = WorkPhoto.objects.all()
    # work_landscape = []
    # work_portrate = []
    # for w in work:
    #     with Image.open(BASE_DIR + w.photo_url) as img:
    #         width, height = img.size
    #         if width / height >= 1.77 and len(work_landscape) < 11:
    #             work_landscape.append(w)
    #         elif width / height < 1 and len(work_portrate) < 11:
    #             work_portrate.append(w)
    categories = Category.objects.all()
    subcategories = SubCategory.objects.all()
    form = OrderForm
    searchform = SearchForm
    socials = Socials.objects.all()
    try:
        logo = Logo.objects.get(inuse=True)
    except:
        logo = ''
    try:
        intro = Introduction.objects.get(inuse=True)
    except:
        intro = ''
    try:
        contacts = Contacts.objects.all()
    except:
        contacts = ''
    allcategory = Category.objects.all()
    allsubcategory = SubCategory.objects.all()

    try:
        data = json.loads(unquote(request.COOKIES.get('cart')))

    except:
        data = []
    ids = []
    for cart_item in data:
        try:
            cart_product = Product.objects.get(vendor_code=cart_item['id'])
            weight = Packprice.objects.get(id=cart_item['weight'])
            cost = weight.price * cart_item['qty']
            ids.append({'product': cart_product, 'qty': cart_item['qty'], 'weight': weight, 'cost': cost})
        except:
            pass
    query = ''
    try:
        query = request.GET.get('text')
    except:
        pass
    search_products = Product.objects.filter(name__iregex=query.lower())
    pageinput = request.GET.get('page')
    page = request.GET.get('page')

    paginator = Paginator(search_products, 12)
    try:
        search_products = paginator.page(page)
    except PageNotAnInteger:
        search_products = paginator.page(1)
    except EmptyPage:
        search_products = paginator.page(paginator.num_pages)

    if page == None or page == 'undefined':
        page = 1

    page = int(page)
    pages = []
    for p in range(paginator.num_pages):
        pages.append(p + 1)
    disc = {
        'pageinput': pageinput,
        'page': page,
        'pages': pages,

        'query': query,
        'search_products': search_products,
        'cart_products': ids,
        'pagename': 'search',
        'allcategory': allcategory,
        'allsubcategory': allsubcategory,
        # 'work': work,
        'contacts': contacts[0],
        'socials': socials,
        'logo': logo,
        'intro': intro,
        # 'work_landscape': work_landscape,
        # 'work_portrate': work_portrate,
        'categories': categories,
        'subcategories': subcategories,
        # 'promoslider': promos,
        # 'services_category': services_category,
        'form': form,
        'searchform': searchform,
    }
    return render(request, 'main/search.html', disc)
